import streamlit as st
import pandas as pd
import numpy as np
import json, re, io, os, base64
from datetime import datetime, date
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="IGSA · Gestão de Prazos",
    page_icon="⚖️",
    layout="wide",
    initial_sidebar_state="expanded"
)

LOGO_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAMCAgMCAgMDAwMEAwMEBQgFBQQEBQoHBwYIDAoMDAsKCwsNDhIQDQ4RDgsLEBYQERMUFRUVDA8XGBYUGBIUFRT/2wBDAQMEBAUEBQkFBQkUDQsNFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBT/wgARCADIAMgDASIAAhEBAxEB/8QAHAABAAICAwEAAAAAAAAAAAAAAAUGAwcCBAgB/8QAGwEBAAIDAQEAAAAAAAAAAAAAAAIDAQQFBgf/2gAMAwEAAhADEAAAAfVIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHUag0NrbLza1O96Sebvh6SyeaZyWp6HRUr1+EGQAAAAAAFM84eoPO/nvXa858L1V9OmtXbIqsOLXrZU9ly3di7doF/73xINmIAAAAAAHW0B6I07zNrz7f6FfeV9ca3lrozrXeGovRsuHfJE9L80DIAAAAAABSrqrz5z1v7Gq/F9X5S7m07ZX2tebykJnpeHDoaIAAAAAAqkbbWqP3F1tV6WzT20LB4sunLoQeY2sruYWJWOcbbI6fSlVMjNYAABx6OJSFCvkZHYo+e25YbkbSr3mOVMunXzXI0+39HNclry3yDNGm5jHi3owF06iM4x9OznyDHkzEADVuzdPbgo7ULT7RRY37Krt11NLWYOGWjvX+Jloq/h1yfg5OvoWHo97LdyIyb1dtiN+mtx9bVcNu1XzU+4p6Wrto6c3GBfxwK73JZG+ATOGGYySy48wqHYtHCvajelYM0qq9ym+RC97tcUKdYZEtzwktwzRx6kpmmgp0zEMwAA4Y86LHxzDB9zDD9yjF9yDEyjD8zjF8zDhzMgyAAAAAAAAAAAAAAAAAAAAAAAAAAA/8QALhAAAQQBAgUDAwMFAAAAAAAABAECAwUGABMREhQVQBYgNQczNiMwYBAhJCZG/9oACAEBAAEFAv5bIVFC7uA+u4D67gPruA+u4D6jNhlk8EohBory6Suj9av160fr1o/XrR+vWr9B2qnJWnssRfAtld1OTukW301Va5zp24i57pF1jG4tti75O5eBcRcY8vg/ydEjdirGvc3DiC2E1esTg5B8SF5B/AJi3oMlg3YxuHU5p97/AIv+gY/RUNeN0YXg5NBsllRbBM1jJcVnWnc9hZKNS146nWEA/Um+FklbIa2xxVpWjagquVSplaJVEm6qa+KqnpYF8ScOIjRdM/hPWRQFi42UZoGoFA8flTm8FllO++q7l5VmRbytvbu3SmEIKjFGEJIMFJuiInlOljhfZEMvdTW7Ybq/s5akOwsJqmOYlsQtPYttq39qH82uYnhkTx7WUWAq2jI1nuMKrCGS1dwS0ogv7U35tKQ2KTIoCFrcrIYXj2UTIykFgkGBx7jW2/vc5GoywGkXii6kFDhcNIE1044iOhMrI9BJXufP0QqyG1k0A8I/ApK6CaEesLWWIdg3bA1ikUACSBopDiO3QEMe2Rrzx43Me2VnsQzv+VyRMljp65a1tCauRXdsC2wAw+6ddVOFfNZQ3lyfJPgMKRFxfIP9TNzRySYpjfwGVslJqMQs+6UVrCwgIWWTBbkzlfDl9tKI0UaMMeQxaHLfZjLFDzKV6xtqsgguV+n8ShSyvSKL6cDPjrMSGYTcvd6ay3JPgMJ/GPqK7nrcrh6fDMb+AZJK8vGnOpMoP+xc1EN0Bj9mRXkZpCsN1rJIVOzH2H0zCynNNeynpB6SCeqRTjAp7KBo/TDU2Pz1Bt9j6209iERYV9TVl1FYPRN6++q5bkAAIoCurh5hR7fHZrK0LHMIij5+Q+mhOMOBhsRYIixogapgpHulfttcRw0kiue0jmcyfmVpPPH1HFZJFYrZ+Z8k6RP3v1EJ4t6j+yktbre/UaRx1E/cZ+w5qPRYWLrbbzbTUXYYmliautpunxtk1tt5nRtdpImo/abw2WcuyzhtNR6QsbprEYn8o//EADsRAAECAwMHCQYGAwAAAAAAAAECAwAEERIhMRMyQYGRofAFBhQiMEJhcdEWICNSseEQMzVAUHOisvH/2gAIAQMBAT8B/fS8uuZcDbYqTHs5P/JvEezs/wDJvETXI03KN5V1N2rsubjyW5xIPeqIfmA0Ut4qVcBHJqlNl3q1+IoeX2jlh0MyLhOkU2wcexl3Sy4HE4g1iaqX5ecRm3/5C4w3lJFVgC1lFk+VdWAjnTMUShjXxv7IXRIc4H5RIbV1kjjGJjnQhKaSzd/j9ompx6cXlHjU9g22pytnRfHRXQso8K6qVjJLshVMcITLrWopGPHpCGVOZvFIQ2pyoT57ITLrWEkd64ca4yarGUpdh7oSTgIYcydsUzhSBNWaJUnQoba02VMNvqYTSzmkKHmPWGV2HMqEE47wRDC8mSkCtQRthl0SziVAX6ePGEPp+GlKSbJJ+npCXico2oVtfWtYCSbgPc5TSG3UNt5oSmmsVrthCFhmZceHxAE7zfx4wylK+T3ivulNNda8eELu5QcUi9QRhoPUH/aQP0s/2D/UwzRfQnlZ5XTzAIp6QVFucU6O6qu+J5oMTlpvNVRQ8j6QunKbfSU/mozvEfN67YmkpRJy9jTUnzrTcIn0pybDneKb9pA3fjllkAG+mEF91Sy4TecYLiiLOiOmv5bpFet9qfSBMOBvI92tdcdJdDiXa3pw8KQt1a6103wZl1QQk93C6EOLbVaQaQHFBNjRC1qcNVfxP//EAC0RAAEDAwICCAcAAAAAAAAAAAEAAhEDEiEEMSIwEBMUIDIzQFEjQUJDcYHB/9oACAECAQE/AfXajU9Vwt3XbKvsu2VfYJusfPEEDIkcmu22sZ6Go7LTiKY5OrZkPR3WyAucGoCBHJezrG2lP01WnncJlNz/AAhUdLY692/IJhXiJUhXAKYRMK4BTmO84TCsRFyOREpwlEXBFu8qNj3aWQSUSJaBsjioIQ8sfn+r7v6R+sLdkKmbmZXlG35JuXulU93DpgI2twpavhxauHxLgtKuaFLBJRc04WDlAR3IBUBWj2UBQFAVoVo9d//EAEMQAAECBAIEBwwIBwEAAAAAAAECAwAEERIhMRMiQVEFFCMyYXGhEDM0QEJSc4GRkrHBIDBicoKy0fAGFUNTYHTCov/aAAgBAQAGPwL/AC21awDHfBHfBHfBHfBHfBFiXAV508SKj6hGlXyjqzgmuceCp9+PBB78eCD348EHvx4Kn34TOM8m+0dZFa0hLqc8lJ3HxGh5tMIdCzUClvV3AUkhWykVJcDtczW7nRcolR3nuICDq0N/VCkt96KSVDxFLm7CEu+o9xojCdmM17UDcIuCiFXZj78JFiUvh3WKRS7DPuTc0d1g/fsh185rNo6h4itG8QobbKw1dzbhWJX7pj8X/fdabprKFT1mGWfNTj1+JDcamHEbjDegdUidZ5zaVUKx0b4s08xd5t5hmXL6uPYFVqtYdZhlo43r1vnEs35IVefV4mypoVKc+qL0PWvfaGBjlWzb56cRFpeXbuuirbdEf3FYJgOlZmHstUUAgvrQpBIoAoUPimsnHeINhvHbHgSFPnKic/wwFTK9Cjdtjk29fz1YnxetMd/iTshydqGQ7fad+WcTsi+hKHGcUFPlpyrEvINpTY4lRLit4hL6k3C9KT0CuJhb7irWkJuJ6IRMBKGwsXJaVnTZUxwaA2hHG12FK8SiKtlF2HOTDXB/J0WyXb7TvpTPuMSBT31tSgv7Q2eyA+2EK10otUN5jTvJS7Kjvim8FI6abYU+NdITcLfKiXmhhpE4jcdv1cz/AKafzRJcKsIK1tPqacSnykKWR8Y4JB52idKuuJpm1KmyjRayqUPs6vZEzKUrOMDQLTtqk/pEs6CLNEPVhHAEyMGlTOfWMI61J+MS3+mr80IQTiuvZCOEENp00s5xkG7Gm7Ld8IYebNUOOtKHtiZb5zjydE2gZqUY4PlMFrabTfU7h+vwjhDgxYCUk8ZZA8059v1BJNANpgBMw2onKihjB6IU+tplCtrigB2xomdEgqxsApWNO82yFD+osDD1wsNPSqfOCFJhapTi5UecWadtIueLLVdbXIFemLVvyq2TsK0lMJW0EEbFDGNJMcWbdpznLQaQVNNyrxpQlIScIUlxDYl0ipChqgQG+Ks6MZJsFIvWZeXXTnKokxxlkNOE/wBVFDX1wHH+LNv7FOWhUBSVBSTtEEF5NU545dcBSFBSTkR9FyUOMjJCpRsWvp/eyChaQpBzBiaRcpSFPFaCpVTSgidml4y8qdGwjZ97r/WHWsl0qhQzSrYRAU74Q0bF9PTH8Q+n/wClxwSZHw4q5SzzajPtjhD0CvhEnXHBX5jEvwjKCyWdXZMMJ5p6QN8Tak4ghBHvCOD/AECfhD8sx3xaCr8Iz+Q9cS6yauN8mvrEFpxNyFrQlQO0XCOLPEq4KmTVCz5MNnBQ0iCPeESkjLK0cxOLsvGaR+zCGWk2oSKAQ3LDCTnxWzYlzo6/n9HhhheCl1WnpF1fnFQ2pzoTT5w4JZt2jZtWpQAp2xwrJuYOtOCvbC1qwSkVMTL6hRLzmr6o/iALr3/YojylxLS7Hgk6MUqxIVlnnujhD0CvhEl+L8xiVlki5518Wp34H9RDzRxKENp/9Jjg/wBAn4Q+6hnSIHJpN1Ms+34RO8GrTo23+UaT2/D4Qn0jf5xC5Z7bilXmnfH8in66VtxOhVvFco4FnT3lDoSo7tYHucDsIxU3RxXQLq/L6LU22sy841zXQK1G4jaItvZbV/cAJ7IU2xcSo3LWo4qMCdl1aGZtsVhVLidxgsOrSyyrBeixUobq7IS1LBDYQKJFMBE8+JhtzjS71CwimJ6emJOZae0MxLKuSSKgw7LaRtGlQUKVaTSGpNt9lWjrrqbO+u+BPTbpm5pIogkUS390Q7KJdQ0hylVFNTnWGpZLrSi0gISooP6wlt5aHCkc5KaV3mJWdRMIYdlzhqVuFcjjCUpdZQQpKiSgnI13xylCr7OUSk0dV+XVUK3jdC5d9F7S8xAavQ/bgHHMFevfD00tWmm3uc4Rs3DcPp1pWMqgCpNYUAMBthApzk3RiKZ9kBSRiTSlYRhzoAArWEgDAi6sISfKi23CttYrbu9u6ObrVtpWG/t9MW0wrbX1RzcxcOmK/U0MDVyi7bAwywEHDOMoThzco1hWLqYxiIupjBFMDiYpTDOKU2Ui6mMGgzig/wAp/8QAKhABAAIBAgQFBAMBAAAAAAAAAQARITFBUWFxgRCRodHxQLHB8CAw4WD/2gAIAQEAAT8h/wCtxwO6nwTPgmfBM+CZ8Ews7CN6fRZenTiMr5wbcz0PafJfafOvafKvafOvafJfaXtlHaD0ZhT/ACEfQvD4BMRmEHYp734HSDk6ZRt7KCnr0jFb1e3wNLHX2a6PeodikFpRp3t+hr7V26MoEbHlY+z4V+yOQ88ByZ6ynb3Gpvmjnw1iNqXPX9fCnmcntbFSv0h3fT6HiRg67THWw6isL9kly/K/KfwGmgob2X/kLdA+4+t/RaA0DvT7zbEx9NoG6vVVZacg11g1ObVLb6XN0yAcq1xOUuQqWdU1Xlcwdofkfevo7WqTlhW0RT8kesaIh8zt3luHRZ1Cbm7hu7+JVQDvFwvK+Up6SrRq/jy+k226Zg06Rge8rG+qV9i43w4ea8jQgok/oXbt9PtrGsM/RY4iOQ3TD7ovJ85V7geZ6wWsLFQBoLNk843F6nfB2D1qFuS5cSq6OHJnIwpya5zBJPlNcRzBhQQ2I2hszIC1zitB4KT3CDA/efLjLv6xXAuxl9+gBKtZFpxyd5okz5aMB1gFmQfRHZH+yzUkKdWHZY6xV4t43WX1ZbHe6JC1Mv8AcXTLB1Ga7ggVzVVoplfSbc6PTJ6oipuI8vhKN4JFuwtft5wSQZlfejo3mLiiPkhJVO3yWNAN+PaUxqrWYjWngYV6k1ZszTB+X9AdjWpQTF+kq7OMGARVTW0q7anSucZIlUT32G/WH8cc9bmmb/cSzzp6y4Orv1YZEbs2Lq+8V7/8M9agKmc2gcmUT5xk8Xa4jhXO32U2j8kCOIs0qLUuTDehU4oRFTqwNieXlMaJ56ChwXMztkHsYM7g1va7wxj2uxOv8Wb2e3aGO4K4gVu0OxImUsK7jtzikzwiU02zLd9o14Q78jww9VsjFzuz8dY7j1HwEPBsLrcTHat+0/e8UNwBtGbdivp6s2lB8jnHmtgbnh1PnnQao2DqwYEl+Y7lPeGi0ySGke6/JX49TA8qYx9lwycSPHD45gpzaesERXD91ji0vOSXwWC/4t9Vbco9DLYm+O9REYqoB+ryg4YGdzC/T1jwWLcAI4TWW4NfNTtDhYezRIooqVLRcHxbl18OvTT6fxvWgffzIzyz3jQeGWFjKvzPMnbNx85YbA7k8IkFKMCZ2xBZcR4JcuFZPKFlyOP8C/LwKmhS2f4nH+Oibn7jjEJQK8yj92F3kRuOf8hHcpeQHGmzqc9IW56U7lYFHfDCDWPwjpKujp1koibLf58ac8T7xqL16DhmGIQDIWtHXAUGATej1VYVxrgSHiOEpfQkNFWkBVAXr9y1yuZbO2xrIPMebAPsBGmldFQgW3qE9UA7F86j5QtRKT7nBlNuMpVtQsXPFzAyBWUND2Plf5rX1PKZKVk2x4cdJnA9Kd6vTvKC02rS+0wlekN7KmHfuTaJiTe1dKQZcJ23rSDjR2cG/aa0CrXT9Ug+gVt6uAgTVKvc1CinkxdXr0mY10aP1SDzVwrb5RKt6WltB8ymxOQ3/TnBXJqKWGAHSZNKetOsHJROB2gLnxl/eptrGZyzgYqUmgmJgBWHaIqCpUSHPvNH7aOLOo27N8bm0a9CCmv3jCppUqDRr/1P/9oADAMBAAIAAwAAABDzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzL3SjzzzzzzzzzoF1Bzzzzzzzzw1MFzzzzzzzzzzwtVXzzzzzzz/AN33Q57188882ldBJSRsM+984AOWQENpWiqf88MZJnqT6wWc9888/fd/vPdN+8888888888888888888888888888//EACYRAQACAgICAQQDAQEAAAAAAAERIQAxQVFhcYEwkaHwECDBQNH/2gAIAQMBAT8Q/wC6fIID947c/X/74Jr9fvEcEoTKie4XNfRFsoHh2feI+ci1J8iCVegLaeovIDAWJXDj7N7j3g7zQHun4mfjHP0WvRD4Zxj0kCugCSmDhdFTk+N1C02W0wW1vIXbZfxR+uvpKU4WxSBpDoH+jlDa2wgfW3uT1juSVwQdAaPoCvhl6kP9MWEs5a5E8188bytpGnkkSfk++RxwmbOBX8Kyq3rCVRQr6Ev4xsMoK9CV+DDrGSE2pEnukdzBLl7QRPlFD7D/AFDVEMMqlUR5R/zFkkhWLQgJ5ewY4ybnbUYNk+KTHIZNGFHoQNWkzxrV0awqIuRGvGMQQkEiS5N6DEYWmNOWalfcxzWGIo2BYbI8WdwuJqFMRGH+af0OEwL2Uy/HBkKgi9gAo4UgfC4c2hSvmAHsBTwcs6BR0FwtzBSFpEn8UCnNB5kZ7iUPRHGCoKgmIUput38YdrRJ7SI7SfGDQQI9VR8msunFnY2egg4lecN3n5Qg+UHuJ/lZ8KyBgmY9TwyW1eLfZxNyPDOzVcVGjAvGzHE9vb79YqdMiYNaEiE01+cRiJUwaCTqdKRrDhRaESSQRADcRE+80PeVG/1a1jsiGJCiZjVk3c+MsUQldIiekUyQWZTDw9nU1MbgnRkp5aPgoDoCgKP6S5LkuS5OTkuS/wDd/8QAKREBAAIBAgQFBQEBAAAAAAAAAQARITFBUWFxgSAwkbHREECh8PHB4f/aAAgBAgEBPxD766Da9pyv5+Z+0/MeQw/ecAtB8lOLZi1iYX1+k9zfPk3C6PuQoGF48WIduwAG3kkkGKdrX0/sTqAEljyBpe8t6ld9JmS9IktiNcDJihHaU5nhsJluDcbWjuPp/IK11Ee/xDkjb8VKA3uPpLouJktQPf5icDp7Sw8CdcW36xDuV/jT95Rodw32qpkTpu3IdPV7kxr0q+6MAB7n+RaNZh7Q1OvTk8PiN4BVdKieAHHp9alTeIr4qVN3lxDotd+vXeW3PJjfepaDZw94aJ0xvtDIKvXXpArZ/wCQMHeAKPA6wnKmGqTFVTaqcqcknJgAUfe//8QAKRABAQACAQMCBgIDAQAAAAAAAREAITFBUWFxgRBAkaHB8CCxMGDR4f/aAAgBAQABPxD/AG3Wylas9jP2z8Z+2fjP2z8Z+2fjP2z8YEsYKAlQeZT5KE/VW34P+4hPNfsqMYE6co+DEZgvwrVJiL8KxfoG+jWfFKiTueclY/KqP2Oo9RPkddDlwDy+tH6GccJpLDA9Veb8D2IqC+E3cQSnAOqrw/p4yXyLLahV8H2+DkbYLeodn6szcd4z1U8AL2X5GBNtfofc++dcZDw1+jr8CKKiGiFPg0b3Dxgm0tR6oN3BZSXBypKNF9+V8NOlfT/dr9GamRK9jaeqHyII1WV2b+4MQG3DN6x6vHvgBFHXGm32xmna+u38Yc878ErI2bDY+QfowRAGR1W36vkmhwvYT8hHtkoKHv7+xMQnkgiAkV5A0DZSj6+jdyXCe0JEKBNIQpviY5qsQtND1YWEjcQNbgPhR9/k6Y7ZtQow2xHjvk1MnVTjRs6Xs4xKd62vbX7IfGaCNtM9pZhWw7P9CfSnxkdkh0mfaVgi6wYiMcNqnSp8pDKKca336+9xyEEGRdo6Hjd7YDTiKd3P/PkyAGl4g6fudsB3iO8939AfLqBuUJQdr218lDd2pq60a57uxhwawlMKMQhK7jpjH4gP0gTl3ldsLcmha5zojOlDrnIdgIUKTnXAc4BK7pNAYYFAkzhVRKwGSWhGJJD1zftq7kQFOV5fzjQ1xBLhp3zbrs/BfAQLSEXi05yXKwlMNlIElsm+5h4uZammTU0QNhGWuY2Dkl5UgTlTHJRs0JR+ie3+SMsdO0keSCO99M4PAYDFPFJ4mKKE0BHAMdKkTCYpQooHVSvKuHwRpCB7FCN4RvGRTsO4gphGNDm9SZ91/wDV9Bx5YfOMIntGUL0jIIg4ahPAEx9XPIQq/tw4XWczyDSLU4E9MvDag88hOk1usv8ABJNYtOCAn+AreAgByq8GLy+jH2h+zJI0KqoMezEfcxE5qtTwgMrwvLl0PmGDUo7tgdtzX0MTqEhS6Od6wsObdyWzyeXnCfrC4pVsvriexgTB2qBTXd5zkiBypcKW/vgl1pXASzD01i3kH7SG0O8XKlT6ScA7Th0zDYznkFSERbdc4nFxTqbHSJDgw87aXazoMXXbCIElCUgtQ2S63g0/AVGtaOyPHLg8qiTdxNOOe6IA/I19LBiQIvcINJ/F/wAkvDF7BC1a65LGivZkdZB5h58NFDDVhyjXy20WOFaVwgd8FNQQiFI3AaeThwezZSQDPooX6Aza3ZmPBJ0SVHuKeg3QYSnXEBaQCicSYsPlGJcA5iTfco0vDyRx+jhLdfw8buH5Q3mGh5Zyy7rWIC8r6ywpPxVCHqLj/OXBQL9QJzBxMRGCkhYD+cJCAXZeeiAPQIjECOftByr1TVXaqvOeRMDST8BGlrp/B4wfXoVR0O+x7PbDc1FBPfSnvc0cM2uoJs4Shx/xbTYAO4wj2HfDapTwgr9DNMIMTXJ8e4PJbvhyVrucdceIg3S4SgCmgjtP3Pd8PSuXPTYIeixKaI9TD9s/a9mXkCMo1QbVZ74O+OvKSoGnZPcHwXS7naCj7SxOop1yvsK2E7qmv0BUgY1qi+ACvn7vDZjOkH1kl7aXunf+JvQTyNEkoLqiXQxbEK0PcAHtSOtziYPVPQgegD+8Po+RQEGpUEUhGMUH27nhAJqERQTnG7kmYRIh48884IS51C1q8ip0MKZdWAnwJBidxN6J4QCRQTPA9eXtjgXV3bkjhnnpiUWOm8grTX3BhAjtHaAKMnu5OKVeaj3WDBOuWfcFbUO0KnVcnng7IcjRqzkyflVTDRGDK3ds6YV7b1sagl484WQe6d/k7U7Pq3hfBBHkTkGInCYOC1qSSKESh8Rim9YzMXXSNVV2j+blLAJUAvKgoHo48mNhBGcO49MAq63XRpGyDdxZSIKtOjSvumXQPt8dg0TkeuCy1QAnZUHpH3w5H7xtyGhu3wa5w1fHOoFejvKrwyImgQE9yZMzlIIoD50PdemKShId085NEu98Yr4Dw1vHaKV84jQA1wRd045cXxjobURJBQosuh73plS1QbXrSaJerxhuid1A0k0sQ3zlwGVFUJ0aCPhP8O0qrsX1N4m6EjRGwThDoOBFkRIGkqcLNXGHF0CDgl2euEAYKVaLU8CvBg546xSCDrxjdGtLsNKa6aNeMLMFwqckeOnjIGhioEWENIVwba2vYVPTbgoLVau5FDowDWdfT7g6+uj6Zq8STX1C3zcB3Upeg2H3wwC1WruShwMA1ggGiRdDtDtXepnNwKqqryq7X/af/9k="

# ── CONSTANTS ──────────────────────────────────────────────────────────────
COORD_MAP = {
    "JULIANA MIRELLA ALVES RODRIGUES": ["JULIANA MIRELLA ALVES RODRIGUES","ARTHUR MASSARI","DANIEL BARROS DE OLIVEIRA","GUSTAVO LOPES ALENCAR FILHO","KELIANE DE OLIVEIRA","MONIQUE DE KAROLIN SILVA DA COSTA","NATALIA PAIVA DE PAULA","ROBERTA RAYANNE VASCONCELOS BOTO","THALLYS ANDERSON FERREIRA DE LIMA","VICTOR EMANOEL FRADIQUE ACCIOLY FONTENELE"],
    "GABRIEL GIORGIO CICCHELERO": ["GABRIEL GIORGIO CICCHELERO","ALYSSON NARBAL DE OLIVEIRA SOMBRA","ANA VITORIA SALES DE OLIVEIRA FALCAO","DALILA DRISANA GOMES GONCALVES","JAMILE BARRETO","JULIANA DE OLIVEIRA ROCHA","RAFAEL CAVALCANTE BARBOSA","RODRIGO RIBEIRO ANTUNES QUARIGUASI"],
    "SUZANA MARIA CAMPOS MARANHAO DE LIMA": ["SUZANA MARIA CAMPOS MARANHAO DE LIMA","SUZANA MARIA CAMPOS MARANHAO DE LIMA","SUZANA MARIA CAMPOS MARANHÃO DE LIMA","EVILANY GABRIELA BRAGA PONTES","FRANCOISE CATHERINE SOUZA ALVES","GIOVANNA CAMPOS PEREIRA","MATHEUS CAVALCANTI DE ARAUJO","TATIANE CARMO SANTA ROSA"],
    "YURI ALVES BARROS DOS SANTOS": ["YURI ALVES BARROS DOS SANTOS","JÚLIA MENEZES MORGADO","JULIA MENEZES MORGADO","LUIZ GUILHERME GONCALVES GIRAO"],
    "NAYANDERSON LUAN MELLO PINHEIRO": ["NAYANDERSON LUAN MELLO PINHEIRO","ANDRE VIANA GARRIDO","EMERSON DE ALMEIDA MELO JUNIOR","EMERSON TRAVASSOS TORQUATO","JEAN VICTOR NUNES SARAIVA"],
    "RONALD FEITOSA AGUIAR FILHO": ["RONALD FEITOSA AGUIAR FILHO","ALEXIA ALENCAR CAPIBARIBE"],
    "LUCIANE MODERNEL MENDES": ["LUCIANE MODERNEL MENDES","ANTONIO EDUARDO GOES AGUIAR FILHO","ERIKA PAULA SANTOS LIMA","SANE BORGES BORGOMONI"],
    "JENIFFER ROSA BARBOSA DE SALES": ["JENIFFER ROSA BARBOSA DE SALES","PAULO MARCIO SOARES DE CARVALHO FILHO"],
    "MARCELLE LEITE RENTROIA": ["MARCELLE LEITE RENTROIA","MARIANA MOTA FROTA","YASMIM GORDIANO BARBOSA"],
    "CAMILLA GOES BARBOSA": ["CAMILLA GOES BARBOSA"],
    "HELANZIA DE ARAUJO XAVIER WICHAMNN": ["HELANZIA DE ARAUJO XAVIER WICHAMNN","HELANZIA DE ARAUJO XAVIER WICHMANN"],
    "TICIANNA PIRES DE SOUZA": ["TICIANNA PIRES DE SOUZA"],
    "TATIANA KOGAN": ["VANESSA NUNES HOLANDA"],
}
COORD_DISPLAY = {"TATIANA KOGAN": "CONTROLADORIA JURÍDICA"}
INACTIVE_SET = {"ANA VITORIA SALES DE OLIVEIRA FALCAO","DALILA DRISANA GOMES GONCALVES","RODRIGO RIBEIRO ANTUNES QUARIGUASI","LUIZ GUILHERME GONCALVES GIRAO","EMERSON DE ALMEIDA MELO JUNIOR","JEAN VICTOR NUNES SARAIVA","YASMIM GORDIANO BARBOSA"}
EXCLUDED_SET = {"TATIANA KOGAN","APARECIDO","CAMILLA GOES BARBOSA","MARIA LAURA MELO ALMEIDA","IRENE FLÁVIA SERENÁRIO","IRENE FLAVIA SERENARIO","IRENE FLÁVIA SERENARIO"}
COLUNAS_ESPERADAS = ["Id","Tipo","Descrição","Conclusão prevista","Responsável processo","Pasta","Status"]
HOLIDAYS_2026 = ["2026-01-01","2026-04-21","2026-05-01","2026-09-07","2026-10-12","2026-11-02","2026-11-15","2026-12-25"]
HOLIDAYS_NP = np.array(HOLIDAYS_2026, dtype="datetime64[D]")

RESP_TO_COORD = {}
for coord, members in COORD_MAP.items():
    for m in members:
        RESP_TO_COORD[m] = coord

CLR_HEADER = "7E1F2D"; CLR_HEADER_TXT = "FFFFFF"
CLR_VERDE = "C6EFCE"; CLR_AMARELO = "FFEB9C"
CLR_ROSA = "FFC7CE"; CLR_VERM = "CC0000"; CLR_VERM_TXT = "FFFFFF"
CLR_LARANJA = "FFCC99"; CLR_GOLD = "CDA736"; CLR_GOLD_TXT = "FFFFFF"
DATA_FILE = "dados_publicados.json"

WINE_COLORS = ["#641828","#7E1F2D","#9B2335","#B83A4C","#C85A6A","#D97A88","#E8A0A8"]
MULTI_COLORS = ["#7E1F2D","#CDA736","#2E9E5B","#1E40AF","#5B21B6","#065F46","#92400E","#701A75","#374151","#9A3412"]

# ── HELPERS ────────────────────────────────────────────────────────────────
def fmt_num(n):
    return f"{n:,}".replace(",",".")

def busdays(d_from, d_to):
    try:
        return int(np.busday_count(np.datetime64(d_from,"D"), np.datetime64(d_to,"D"), holidays=HOLIDAYS_NP))
    except: return None

def parse_date(val):
    if val is None: return None
    if isinstance(val, (datetime, date)): return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    for fmt in ["%d/%m/%Y","%Y-%m-%d","%d.%m.%Y"]:
        try: return datetime.strptime(s[:10], fmt).date()
        except: pass
    return None

def extract_fatal(desc):
    if not isinstance(desc, str): return None
    m = re.search(r"FATAL[\\s:]+(\d{2})/(\d{2})/(\d{4})", desc, re.IGNORECASE)
    if m:
        try: return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except: pass
    return None

def extract_aud(desc):
    if not isinstance(desc, str): return None
    m = re.search(r"AUD[\\s:]+(\d{2})/(\d{2})/(\d{4})", desc, re.IGNORECASE)
    if m:
        try: return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except: pass
    return None

def check_incons(tipo, desc, conclusao, fatal, aud):
    issues = []
    d = desc or ""
    has_elab = bool(re.search(r"ELABORAR", d, re.IGNORECASE))
    ref = fatal or aud
    if ref and conclusao and conclusao > ref:
        issues.append(f"Conclusão posterior à data da descrição ({ref.strftime('%d/%m/%Y')})")
    if re.search(r"\d{5,}", d): issues.append("Ano inválido na descrição")
    if tipo == "Prazo" and re.search(r"AUDIÊNCIA DE CONCILIAÇÃO", d, re.IGNORECASE):
        issues.append("Tipo Prazo com descrição de Audiência")
    if tipo == "Audiência" and re.search(r"PRAZO.*Protocolar", d, re.IGNORECASE):
        issues.append("Tipo Audiência com descrição de Prazo")
    if tipo == "Diversos" and not has_elab:
        if re.search(r"AUDIÊNCIA DE CONCILIAÇÃO", d, re.IGNORECASE) or re.search(r"PRAZO.*Protocolar", d, re.IGNORECASE):
            issues.append("Tipo Diversos com descrição de Audiência/Prazo")
    return "; ".join(issues)

def get_row_color(du, incons, tipo):
    if incons: return CLR_LARANJA, "000000"
    if tipo in ("AUDIÊNCIA DE JULGAMENTO","ACOMPANHAR JULGAMENTO"): return CLR_VERDE, "000000"
    if du is None: return None, None
    if du < 0: return CLR_VERM, CLR_VERM_TXT
    if du == 0: return CLR_ROSA, "000000"
    if du == 1: return CLR_AMARELO, "000000"
    return CLR_VERDE, "000000"

# ── DATA PROCESSING ────────────────────────────────────────────────────────
def processar_planilha(uploaded_file, today):
    df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)
    header_idx = 1
    for i in range(min(10, len(df_raw))):
        row = df_raw.iloc[i].tolist()
        if any("Conclusão" in str(c) or "Tipo" in str(c) for c in row):
            header_idx = i; break
    headers = [str(h).strip() for h in df_raw.iloc[header_idx].tolist()]
    df = df_raw.iloc[header_idx+1:].copy()
    df.columns = headers
    df = df.drop_duplicates().reset_index(drop=True)
    return df, headers

def validar_estrutura(headers):
    return [c for c in COLUNAS_ESPERADAS if c not in headers]

def construir_registros(df, today):
    registros, alertas, seen = [], [], set()
    for _, row in df.iterrows():
        conclusao = parse_date(row.get("Conclusão prevista"))
        if not conclusao: continue
        du = busdays(today, conclusao)
        if du is None or du > 1: continue
        id_ = str(row.get("Id","")).strip()
        pasta = str(row.get("Pasta","")).strip()
        processo = pasta if pasta and pasta not in ("","nan","None") else id_
        tipo = str(row.get("Tipo","")).strip()
        desc = str(row.get("Descrição","")).strip()
        resp = str(row.get("Responsável processo","")).strip()
        status = str(row.get("Status","")).strip()
        key = f"{id_}|{conclusao}|{tipo}|{resp}"
        if key in seen: continue
        seen.add(key)
        coord_raw = RESP_TO_COORD.get(resp, "SEM COORDENADOR")
        coord_display = COORD_DISPLAY.get(coord_raw, coord_raw)
        fatal = extract_fatal(desc)
        aud = extract_aud(desc)
        incons = check_incons(tipo, desc, conclusao, fatal, aud)
        inativo = resp in INACTIVE_SET
        if incons:
            alertas.append({"Id":id_,"Processo":processo,"Tipo":tipo,
                "Responsável":resp or "(Sem responsável)","Coordenador":coord_display,
                "Conclusão":conclusao.strftime("%d/%m/%Y"),"DU":du,"Inconsistência":incons})
        registros.append({"id":id_,"processo":processo,"tipo":tipo,"desc":desc[:300],
            "status":status,"resp":resp or "(Sem responsável)",
            "coord":coord_raw,"coord_display":coord_display,
            "conclusao":conclusao.strftime("%d/%m/%Y"),"conclusao_iso":conclusao.isoformat(),
            "du":du,"inativo":inativo,"incons":incons})
    registros.sort(key=lambda r: (r["conclusao_iso"], r["resp"], r["processo"]))
    return registros, alertas

# ── LOAD/SAVE ──────────────────────────────────────────────────────────────
def load_published():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, encoding="utf-8") as f: return json.load(f)
        except: pass
    return {"registros":[],"publicado_em":None,"total":0,"versao":None}

def save_published(registros, versao, today_str):
    data = {"registros":registros,"publicado_em":datetime.now().strftime("%d/%m/%Y %H:%M"),
            "total":len(registros),"versao":versao,"referencia":today_str}
    with open(DATA_FILE,"w",encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, default=str)
    return data

def push_to_github(token, repo_name, file_path, content, commit_msg):
    try:
        from github import Github
        g = Github(token)
        repo = g.get_repo(repo_name)
        try:
            existing = repo.get_contents(file_path)
            repo.update_file(file_path, commit_msg, content, existing.sha)
        except: repo.create_file(file_path, commit_msg, content)
        return True, None
    except Exception as e: return False, str(e)

# ── EXCEL ──────────────────────────────────────────────────────────────────
def make_header_fill(): return PatternFill("solid", fgColor=CLR_HEADER)
def make_header_font(): return Font(bold=True, color=CLR_HEADER_TXT, size=10)
def make_header_align(): return Alignment(horizontal="center", vertical="center", wrap_text=True)

def apply_row_style(ws, row_num, du, incons, tipo, num_cols):
    bg, txt = get_row_color(du, incons, tipo)
    if not bg: return
    fill = PatternFill("solid", fgColor=bg)
    font = Font(color=txt, size=10)
    for col in range(1, num_cols+1):
        ws.cell(row=row_num, column=col).fill = fill
        ws.cell(row=row_num, column=col).font = font

COLS_DETAIL = ["Processo","Tipo","Descrição","Coordenador","Responsável","Status","Conclusão Prevista","Dias Úteis","Inconsistência"]
COLS_RESUMO = ["Responsável","Status","Prazo","Audiência","Diversos","Workflow","Publicação","Outros","Total","Inconsistências"]

def write_sheet(ws, data_rows, cols, is_resumo=False):
    hfill = make_header_fill(); hfont = make_header_font(); halign = make_header_align()
    ws.row_dimensions[1].height = 30
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hfill; cell.font = hfont; cell.alignment = halign
    for ri, row in enumerate(data_rows, 2):
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col,""))
            cell.alignment = Alignment(vertical="top", wrap_text=(ci==3))
            cell.font = Font(size=10)
        if not is_resumo:
            apply_row_style(ws, ri, row.get("Dias Úteis"), row.get("Inconsistência",""), row.get("Tipo",""), len(cols))
        else:
            if row.get("Responsável") == "TOTAL":
                for ci in range(1, len(cols)+1):
                    ws.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=CLR_GOLD)
                    ws.cell(row=ri, column=ci).font = Font(bold=True, color=CLR_GOLD_TXT, size=10)
    for ci, col in enumerate(cols, 1):
        max_len = max([len(str(col))] + [len(str(r.get(col,""))) for r in data_rows[:50]])
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len+2, 8), 50)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

def gerar_excel_coord(coord_key, registros, coord_display):
    rows = [r for r in registros if r["coord"]==coord_key and r["resp"] not in EXCLUDED_SET]
    rows.sort(key=lambda r: (r["conclusao_iso"], r["resp"], r["processo"]))
    wb = openpyxl.Workbook()

    by_resp = {}
    for r in rows:
        rp = r["resp"]
        if rp not in by_resp:
            by_resp[rp] = {"Responsável":rp,"Status":"Inativo" if r["inativo"] else "Ativo",
                           "Prazo":0,"Audiência":0,"Diversos":0,"Workflow":0,"Publicação":0,"Outros":0,"Total":0,"Inconsistências":0}
        t = r["tipo"]
        if t in ("Prazo","Audiência","Diversos","Workflow","Publicação"): by_resp[rp][t] += 1
        else: by_resp[rp]["Outros"] += 1
        by_resp[rp]["Total"] += 1
        if r["incons"]: by_resp[rp]["Inconsistências"] += 1
    resumo_rows = sorted(by_resp.values(), key=lambda x: -x["Total"])
    tot = {c: sum(r.get(c,0) for r in resumo_rows if isinstance(r.get(c,0),int))
           for c in ["Prazo","Audiência","Diversos","Workflow","Publicação","Outros","Total","Inconsistências"]}
    tot["Responsável"] = "TOTAL"; tot["Status"] = ""
    resumo_rows.append(tot)

    ws_res = wb.active; ws_res.title = "Resumo"
    write_sheet(ws_res, resumo_rows, COLS_RESUMO, is_resumo=True)

    def to_detail(r):
        return {"Processo":r["processo"],"Tipo":r["tipo"],"Descrição":r["desc"],
                "Coordenador":r["coord_display"],"Responsável":r["resp"],"Status":r["status"],
                "Conclusão Prevista":r["conclusao"],"Dias Úteis":r["du"],"Inconsistência":r["incons"]}

    for tipo_key, nome_aba in [("Prazo","Prazos"),("Audiência","Audiências"),("Diversos","Diversos")]:
        ws = wb.create_sheet(nome_aba)
        write_sheet(ws, [to_detail(r) for r in rows if r["tipo"]==tipo_key], COLS_DETAIL)

    ws_inc = wb.create_sheet("Inconsistências")
    write_sheet(ws_inc, [to_detail(r) for r in rows if r["incons"]], COLS_DETAIL)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

def exportar_xlsx_filtrado(rows):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Dados"
    hfill = make_header_fill(); hfont = make_header_font(); halign = make_header_align()
    ws.row_dimensions[1].height = 28
    for ci, col in enumerate(COLS_DETAIL, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill=hfill; c.font=hfont; c.alignment=halign
    for ri, r in enumerate(rows, 2):
        vals = [r.get("processo",""),r.get("tipo",""),r.get("desc",""),
                r.get("coord_display",r.get("coord","")),r.get("resp",""),r.get("status",""),
                r.get("conclusao",""),r.get("du",""),r.get("incons","")]
        for ci,v in enumerate(vals,1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font=Font(size=10); cell.alignment=Alignment(vertical="top",wrap_text=(ci==3))
        apply_row_style(ws, ri, r.get("du"), r.get("incons",""), r.get("tipo",""), len(COLS_DETAIL))
    for ci in range(1, len(COLS_DETAIL)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 20
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

def exportar_csv(rows):
    lines = [";".join(COLS_DETAIL)]
    for r in rows:
        vals = [r.get("processo",""),r.get("tipo",""),r.get("desc","").replace(";",","),
                r.get("coord_display",r.get("coord","")),r.get("resp",""),r.get("status",""),
                r.get("conclusao",""),str(r.get("du","")),r.get("incons","")]
        lines.append(";".join(f'"{v}"' for v in vals))
    return ("\uFEFF" + "\n".join(lines)).encode("utf-8")

# ── CSS ────────────────────────────────────────────────────────────────────
def inject_css():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Libre+Franklin:wght@400;500;600;700&family=Cormorant+Garamond:wght@600&display=swap');
    html,body,[class*="css"]{font-family:'Libre Franklin',sans-serif}
    .main-header{background:linear-gradient(135deg,#641828 0%,#7E1F2D 55%,#9B2335 100%);
        padding:18px 28px;border-radius:12px;margin-bottom:24px;
        display:flex;align-items:center;gap:20px;
        box-shadow:0 4px 16px rgba(100,24,40,.35)}
    .main-header img{height:72px;border-radius:8px;background:white;padding:4px}
    .header-text h1{color:white;font-family:'Cormorant Garamond',serif;
        font-size:28px;font-weight:600;margin:0;letter-spacing:.04em}
    .header-text p{color:rgba(255,255,255,.7);font-size:11px;
        letter-spacing:.12em;text-transform:uppercase;margin:2px 0 0}
    .metric-card{background:white;border:1px solid #E5DAC7;border-radius:12px;
        padding:16px 18px;text-align:center;position:relative;overflow:hidden;
        box-shadow:0 1px 4px rgba(0,0,0,.04);margin-bottom:4px}
    .metric-card::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:#7E1F2D}
    .metric-card.gold::before{background:#CDA736}
    .metric-card.green::before{background:#2E9E5B}
    .metric-card.rose::before{background:#D9737E}
    .metric-kicker{font-size:10px;letter-spacing:.12em;text-transform:uppercase;color:#A07B1E;font-weight:600;margin-bottom:4px}
    .metric-value{font-size:30px;font-weight:700;color:#2A2420;line-height:1.1}
    .metric-label{font-size:11px;color:#888;margin-top:2px}
    .section-title{font-size:11px;letter-spacing:.1em;text-transform:uppercase;
        color:#7E1F2D;font-weight:600;border-bottom:1px solid #E5DAC7;
        padding-bottom:8px;margin:20px 0 12px}
    .published-banner{background:linear-gradient(90deg,#1a4731,#2E9E5B);
        color:white;padding:10px 18px;border-radius:8px;margin-bottom:16px;font-size:12px}
    div[data-testid="stSidebarContent"]{background:#FAF8F3}
    .stDownloadButton>button{background:#7E1F2D!important;color:white!important;
        border:none!important;font-weight:600!important}
    .stDownloadButton>button:hover{background:#641828!important}
    </style>
    """, unsafe_allow_html=True)

def render_header(subtitle=""):
    st.markdown(f"""
    <div class="main-header">
        <img src="data:image/jpeg;base64,{LOGO_B64}" alt="IGSA"/>
        <div class="header-text">
            <h1>IMACULADA GORDIANO</h1>
            <p>Controladoria Jurídica · Gestão de Prazos Preclusivos{" · "+subtitle if subtitle else ""}</p>
        </div>
    </div>""", unsafe_allow_html=True)

def card(kicker, value, label, color=""):
    return f'''<div class="metric-card {color}">
        <div class="metric-kicker">{kicker}</div>
        <div class="metric-value">{value}</div>
        <div class="metric-label">{label}</div>
    </div>'''

# ── CHARTS ─────────────────────────────────────────────────────────────────
def chart_bar_v(df_data, x_col, y_col, title, colors=None):
    """Vertical bar chart using Altair-style via st.bar_chart with custom colors via plotly workaround"""
    if df_data.empty: return
    import altair as alt
    color_list = colors or WINE_COLORS
    n = len(df_data)
    color_map = {row[x_col]: color_list[i % len(color_list)] for i, row in df_data.iterrows()}
    chart = alt.Chart(df_data).mark_bar(cornerRadiusTopLeft=4,cornerRadiusTopRight=4).encode(
        x=alt.X(f"{x_col}:N", sort="-y", axis=alt.Axis(labelAngle=-35,labelLimit=120,labelFontSize=11)),
        y=alt.Y(f"{y_col}:Q", axis=alt.Axis(labelFontSize=11)),
        color=alt.Color(f"{x_col}:N", scale=alt.Scale(domain=list(color_map.keys()), range=list(color_map.values())), legend=None),
        tooltip=[x_col, y_col]
    ).properties(height=280).configure_axis(grid=True,gridColor="#EDE5D4").configure_view(strokeWidth=0)
    st.altair_chart(chart, use_container_width=True)

def chart_bar_h(df_data, x_col, y_col, title, color="#7E1F2D"):
    if df_data.empty: return
    import altair as alt
    chart = alt.Chart(df_data).mark_bar(color=color,cornerRadiusTopRight=4,cornerRadiusBottomRight=4).encode(
        y=alt.Y(f"{y_col}:N", sort="-x", axis=alt.Axis(labelLimit=200,labelFontSize=11)),
        x=alt.X(f"{x_col}:Q", axis=alt.Axis(labelFontSize=11)),
        tooltip=[y_col, x_col]
    ).properties(height=320).configure_axis(grid=True,gridColor="#EDE5D4").configure_view(strokeWidth=0)
    st.altair_chart(chart, use_container_width=True)

def chart_donut(labels, values, title):
    if not values: return
    import altair as alt
    df_d = pd.DataFrame({"Tipo":labels,"Qtd":values})
    total = sum(values)
    base = alt.Chart(df_d).encode(
        theta=alt.Theta("Qtd:Q",stack=True),
        color=alt.Color("Tipo:N", scale=alt.Scale(range=MULTI_COLORS),
                        legend=alt.Legend(orient="bottom",labelFontSize=11,titleFontSize=11)),
        tooltip=["Tipo","Qtd"]
    )
    pie = base.mark_arc(innerRadius=70, outerRadius=120)
    text = alt.Chart(pd.DataFrame({"label":[f"Total\n{fmt_num(total)}"]})).mark_text(
        size=14, fontWeight="bold", color="#2A2420"
    ).encode(text="label:N")
    st.altair_chart((pie+text).properties(height=300).configure_view(strokeWidth=0), use_container_width=True)

# ── SIDEBAR ────────────────────────────────────────────────────────────────
def render_sidebar():
    with st.sidebar:
        st.markdown(f'''<div style="text-align:center;padding:12px 0 20px">
            <img src="data:image/jpeg;base64,{LOGO_B64}" style="height:64px;border-radius:8px;background:white;padding:4px"/>
            <div style="font-size:10px;letter-spacing:.1em;text-transform:uppercase;color:#7E1F2D;margin-top:8px;font-weight:600">GESTÃO DE PRAZOS</div>
        </div>''', unsafe_allow_html=True)
        st.divider()
        page = st.radio("NAV", [
            "📊 Visão Geral","👥 Por Coordenador","👤 Por Responsável",
            "🔍 Auditoria","📥 Exportação","⚙️ Área Administrativa",
        ], label_visibility="collapsed")
        st.divider()
        pub = load_published()
        if pub.get("publicado_em"):
            st.markdown(f'''<div style="font-size:11px;color:#666;padding:8px">
                <div style="font-weight:600;color:#2E9E5B">● Dados publicados</div>
                <div>{pub["publicado_em"]}</div>
                <div>{fmt_num(pub["total"])} registros</div>
                {f'<div>Ref: {pub.get("referencia","")}</div>' if pub.get("referencia") else ''}
            </div>''', unsafe_allow_html=True)
        else:
            st.markdown('<div style="font-size:11px;color:#aaa;padding:8px">● Sem dados publicados</div>', unsafe_allow_html=True)
    return page

# ── PAGES ──────────────────────────────────────────────────────────────────
def get_df_filters(registros, prefix):
    df = pd.DataFrame(registros) if registros else pd.DataFrame()
    if df.empty: return df
    df["conclusao_dt"] = pd.to_datetime(df["conclusao_iso"])
    # Filter out nan/empty resp
    df = df[df["resp"].notna() & (df["resp"] != "") & (df["resp"] != "nan")]
    return df

def apply_common_filters(df, prefix):
    if df.empty: return df
    dt_ini = st.session_state.get(f"{prefix}_ini")
    dt_fim = st.session_state.get(f"{prefix}_fim")
    if dt_ini: df = df[df["conclusao_dt"] >= pd.Timestamp(dt_ini)]
    if dt_fim: df = df[df["conclusao_dt"] <= pd.Timestamp(dt_fim)]
    return df

def page_geral(registros):
    render_header("Visão Geral")
    df = get_df_filters(registros, "g")
    if df.empty: st.info("Nenhum dado disponível. Publique uma planilha na Área Administrativa."); return

    c1,c2,c3,c4 = st.columns(4)
    with c1: dt_ini = st.date_input("Data Início", value=None, key="g_ini")
    with c2: dt_fim = st.date_input("Data Fim", value=None, key="g_fim")
    with c3:
        coords = ["Todos"] + sorted(df["coord_display"].dropna().unique().tolist())
        coord_f = st.selectbox("Coordenador", coords, key="g_coord")
    with c4:
        tipos = ["Todos"] + sorted(df["tipo"].dropna().unique().tolist())
        tipo_f = st.selectbox("Tipo", tipos, key="g_tipo")

    if dt_ini: df = df[df["conclusao_dt"] >= pd.Timestamp(dt_ini)]
    if dt_fim: df = df[df["conclusao_dt"] <= pd.Timestamp(dt_fim)]
    if coord_f != "Todos": df = df[df["coord_display"] == coord_f]
    if tipo_f != "Todos": df = df[df["tipo"] == tipo_f]
    active = df[~df["resp"].isin(EXCLUDED_SET)]

    st.markdown('<div class="section-title">Resumo</div>', unsafe_allow_html=True)
    cols = st.columns(8)
    cards_data = [
        ("Total",fmt_num(len(active)),"Atividades até D-1",""),
        ("Prazos",fmt_num(len(active[active.tipo=="Prazo"])),"Tipo Prazo","rose"),
        ("Audiências",fmt_num(len(active[active.tipo=="Audiência"])),"Tipo Audiência",""),
        ("Diversos",fmt_num(len(active[active.tipo=="Diversos"])),"Tipo Diversos","gold"),
        ("Inconsistências",fmt_num(len(active[active.incons!=""])),"Com alerta","rose"),
        ("Coordenadores",fmt_num(active["coord"].nunique()),"Com pendências",""),
        ("Ativos",fmt_num(active[~active.inativo]["resp"].nunique()),"Responsáveis","green"),
        ("Inativos",fmt_num(active[active.inativo]["resp"].nunique()),"Responsáveis",""),
    ]
    for i,(k,v,l,c) in enumerate(cards_data):
        with cols[i]: st.markdown(card(k,v,l,c), unsafe_allow_html=True)

    st.markdown('<div class="section-title">Distribuição por Coordenador</div>', unsafe_allow_html=True)
    by_coord = active[~active.inativo].groupby("coord_display").size().reset_index(name="Pendências")
    by_coord = by_coord.sort_values("Pendências", ascending=False)
    by_coord.columns = ["Coordenador","Pendências"]
    chart_bar_v(by_coord, "Coordenador", "Pendências", "Por Coordenador")

    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div class="section-title">Por Tipo</div>', unsafe_allow_html=True)
        by_tipo = active.groupby("tipo").size().reset_index(name="Qtd").sort_values("Qtd", ascending=False)
        chart_donut(by_tipo["tipo"].tolist(), by_tipo["Qtd"].tolist(), "Por Tipo")
    with c2:
        st.markdown('<div class="section-title">Top 10 Responsáveis</div>', unsafe_allow_html=True)
        top10 = (active[~active.inativo & (active.resp!="(Sem responsável)")]
                 .groupby("resp").size().reset_index(name="Qtd")
                 .sort_values("Qtd", ascending=False).head(10))
        top10.columns = ["Responsável","Qtd"]
        chart_bar_h(top10, "Qtd", "Responsável", "Top 10", "#7E1F2D")

def page_coordenador(registros):
    render_header("Por Coordenador")
    df = get_df_filters(registros, "c")
    if df.empty: st.info("Nenhum dado disponível."); return

    c1,c2,c3,c4,c5,c6 = st.columns(6)
    with c1: dt_ini = st.date_input("Data Início", value=None, key="c_ini")
    with c2: dt_fim = st.date_input("Data Fim", value=None, key="c_fim")
    with c3:
        coords = ["Todos"] + sorted(df["coord_display"].dropna().unique().tolist())
        coord_f = st.selectbox("Coordenador", coords, key="c_coord")
    with c4:
        resps = ["Todos"] + sorted(df["resp"].dropna().unique().tolist())
        resp_f = st.selectbox("Responsável", resps, key="c_resp")
    with c5:
        tipos = ["Todos"] + sorted(df["tipo"].dropna().unique().tolist())
        tipo_f = st.selectbox("Tipo", tipos, key="c_tipo")
    with c6: proc_f = st.text_input("Processo", key="c_proc")

    if dt_ini: df = df[df["conclusao_dt"] >= pd.Timestamp(dt_ini)]
    if dt_fim: df = df[df["conclusao_dt"] <= pd.Timestamp(dt_fim)]
    if coord_f != "Todos": df = df[df["coord_display"] == coord_f]
    if resp_f != "Todos": df = df[df["resp"] == resp_f]
    if tipo_f != "Todos": df = df[df["tipo"] == tipo_f]
    if proc_f: df = df[df["processo"].str.contains(proc_f, case=False, na=False)]
    active = df[~df["resp"].isin(EXCLUDED_SET)]

    cols = st.columns(5)
    for i,(k,v,l,c) in enumerate([
        ("Total",fmt_num(len(active)),"Filtrados",""),
        ("Prazos",fmt_num(len(active[active.tipo=="Prazo"])),"Tipo Prazo","rose"),
        ("Audiências",fmt_num(len(active[active.tipo=="Audiência"])),"Tipo Audiência",""),
        ("Diversos",fmt_num(len(active[active.tipo=="Diversos"])),"Tipo Diversos","gold"),
        ("Inconsistências",fmt_num(len(active[active.incons!=""])),"Com alertas","rose"),
    ]):
        with cols[i]: st.markdown(card(k,v,l,c), unsafe_allow_html=True)

    st.markdown('<div class="section-title">Resumo por Responsável</div>', unsafe_allow_html=True)
    # Safe groupby avoiding KeyError
    if active.empty:
        st.info("Nenhum registro para os filtros selecionados.")
    else:
        resumo_rows = []
        for resp_name, grp in active.groupby("resp", sort=False):
            resumo_rows.append({
                "Responsável": resp_name,
                "Status": "Inativo" if grp["inativo"].any() else "Ativo",
                "Prazo": int((grp["tipo"]=="Prazo").sum()),
                "Audiência": int((grp["tipo"]=="Audiência").sum()),
                "Diversos": int((grp["tipo"]=="Diversos").sum()),
                "Workflow": int((grp["tipo"]=="Workflow").sum()),
                "Publicação": int((grp["tipo"]=="Publicação").sum()),
                "Outros": int((~grp["tipo"].isin(["Prazo","Audiência","Diversos","Workflow","Publicação"])).sum()),
                "Total": len(grp),
                "Inconsistências": int((grp["incons"]!="").sum()),
            })
        resumo_rows.sort(key=lambda x: -x["Total"])
        st.dataframe(pd.DataFrame(resumo_rows), use_container_width=True, hide_index=True)

    st.markdown('<div class="section-title">Detalhamento</div>', unsafe_allow_html=True)
    det = active[["processo","tipo","desc","coord_display","resp","conclusao","du","incons"]].copy()
    det.columns = ["Processo","Tipo","Descrição","Coordenador","Responsável","Conclusão","DU","Inconsistência"]
    st.dataframe(det, use_container_width=True, hide_index=True, height=380)

    buf = exportar_xlsx_filtrado(active.to_dict("records"))
    st.download_button("📥 Exportar Excel", buf, "IGSA_Filtrado.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

def page_responsavel(registros):
    render_header("Por Responsável")
    df = get_df_filters(registros, "r")
    if df.empty: st.info("Nenhum dado disponível."); return

    c1,c2,c3,c4 = st.columns(4)
    with c1: dt_ini = st.date_input("Data Início", value=None, key="r_ini")
    with c2: dt_fim = st.date_input("Data Fim", value=None, key="r_fim")
    with c3:
        resps = ["Selecionar..."] + sorted(df["resp"].dropna().unique().tolist())
        resp_f = st.selectbox("Responsável", resps, key="r_resp")
    with c4:
        tipos = ["Todos"] + sorted(df["tipo"].dropna().unique().tolist())
        tipo_f = st.selectbox("Tipo", tipos, key="r_tipo")
    proc_f = st.text_input("Buscar processo", key="r_proc")

    if dt_ini: df = df[df["conclusao_dt"] >= pd.Timestamp(dt_ini)]
    if dt_fim: df = df[df["conclusao_dt"] <= pd.Timestamp(dt_fim)]
    if resp_f != "Selecionar...": df = df[df["resp"] == resp_f]
    if tipo_f != "Todos": df = df[df["tipo"] == tipo_f]
    if proc_f: df = df[df["processo"].str.contains(proc_f, case=False, na=False)]

    if resp_f == "Selecionar...":
        st.info("Selecione um responsável para ver os indicadores individuais.")
        return

    total = len(df); incons = int((df.incons != "").sum())
    oldest = df["conclusao_iso"].min() if not df.empty else None
    aud_df = df[(df.tipo=="Audiência") & (df.du >= 0)].sort_values("du")
    next_aud = aud_df.iloc[0] if not aud_df.empty else None

    cols = st.columns(4)
    with cols[0]: st.markdown(card("Total Pendente", fmt_num(total), "Atividades até D-1"), unsafe_allow_html=True)
    with cols[1]:
        v = oldest.split("-") if oldest else None
        st.markdown(card("Mais Antiga", f"{v[2]}/{v[1]}/{v[0]}" if v else "—", "Conclusão mais antiga"), unsafe_allow_html=True)
    with cols[2]:
        st.markdown(card("Próxima Audiência",
            next_aud["conclusao"] if next_aud is not None else "—",
            f"DU: {next_aud['du']}" if next_aud is not None else "Sem audiências","gold"), unsafe_allow_html=True)
    with cols[3]:
        st.markdown(card("Inconsistências", fmt_num(incons),
            "Com alertas" if incons else "Sem alertas","rose" if incons else "green"), unsafe_allow_html=True)

    st.markdown('<div class="section-title">Atividades</div>', unsafe_allow_html=True)
    det = df[["processo","tipo","desc","coord_display","resp","conclusao","du","incons"]].copy()
    det.columns = ["Processo","Tipo","Descrição","Coordenador","Responsável","Conclusão","DU","Inconsistência"]
    st.dataframe(det, use_container_width=True, hide_index=True, height=450)

    buf = exportar_xlsx_filtrado(df.to_dict("records"))
    st.download_button("📥 Exportar Excel", buf, f"IGSA_{resp_f[:30]}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

def page_auditoria(registros):
    render_header("Auditoria")
    df = get_df_filters(registros, "a")
    if df.empty: st.info("Nenhum dado disponível."); return
    df_inc_all = df[df["incons"] != ""]

    c1,c2,c3,c4 = st.columns(4)
    with c1: dt_ini = st.date_input("Data Início", value=None, key="a_ini")
    with c2: dt_fim = st.date_input("Data Fim", value=None, key="a_fim")
    with c3:
        coords = ["Todos"] + sorted(df["coord_display"].dropna().unique().tolist())
        coord_f = st.selectbox("Coordenador", coords, key="a_coord")
    with c4:
        resps = ["Todos"] + sorted(df["resp"].dropna().unique().tolist())
        resp_f = st.selectbox("Responsável", resps, key="a_resp")

    inc_types = sorted(set(t.strip() for r in df_inc_all.itertuples() for t in r.incons.split(";")))
    ti_f = st.selectbox("Tipo de Inconsistência", ["Todas"] + inc_types, key="a_tipo")

    df_inc = df_inc_all.copy()
    if dt_ini: df_inc = df_inc[df_inc["conclusao_dt"] >= pd.Timestamp(dt_ini)]
    if dt_fim: df_inc = df_inc[df_inc["conclusao_dt"] <= pd.Timestamp(dt_fim)]
    if coord_f != "Todos": df_inc = df_inc[df_inc["coord_display"] == coord_f]
    if resp_f != "Todos": df_inc = df_inc[df_inc["resp"] == resp_f]
    if ti_f != "Todas": df_inc = df_inc[df_inc["incons"].str.contains(ti_f, na=False)]

    types_count = {}
    for r in df_inc.itertuples():
        for t in r.incons.split(";"):
            t=t.strip()
            if t: types_count[t]=types_count.get(t,0)+1

    ncols = min(len(types_count)+1, 5)
    cols = st.columns(ncols)
    with cols[0]: st.markdown(card("Total",fmt_num(len(df_inc)),"Inconsistências","rose"), unsafe_allow_html=True)
    for i,(t,v) in enumerate(list(types_count.items())[:ncols-1],1):
        with cols[i]: st.markdown(card(t[:35],fmt_num(v),"ocorrências","gold"), unsafe_allow_html=True)

    st.markdown('<div class="section-title">Registros com Inconsistências</div>', unsafe_allow_html=True)
    det = df_inc[["processo","tipo","desc","coord_display","resp","conclusao","du","incons"]].copy()
    det.columns = ["Processo","Tipo","Descrição","Coordenador","Responsável","Conclusão","DU","Inconsistência"]
    st.dataframe(det, use_container_width=True, hide_index=True, height=450)

    buf = exportar_xlsx_filtrado(df_inc.to_dict("records"))
    st.download_button("📥 Exportar Inconsistências", buf, "IGSA_Auditoria.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

def page_exportacao(registros):
    render_header("Exportação")
    if not registros: st.info("Nenhum dado disponível."); return

    c1,c2 = st.columns(2)
    with c1: dt_ini = st.date_input("Data Início", value=None, key="e_ini")
    with c2: dt_fim = st.date_input("Data Fim", value=None, key="e_fim")

    rows = registros
    if dt_ini: rows = [r for r in rows if r["conclusao_iso"] >= dt_ini.isoformat()]
    if dt_fim: rows = [r for r in rows if r["conclusao_iso"] <= dt_fim.isoformat()]
    active = [r for r in rows if r["resp"] not in EXCLUDED_SET and r["resp"] not in ("","nan","(Sem responsável)")]

    st.markdown('<div class="section-title">Exportar por Coordenador</div>', unsafe_allow_html=True)
    coords = sorted(set(r["coord"] for r in active))
    cols_grid = st.columns(3)
    for i, coord_key in enumerate(coords):
        coord_rows = [r for r in active if r["coord"]==coord_key]
        coord_display = COORD_DISPLAY.get(coord_key, coord_key)
        n_p = sum(1 for r in coord_rows if r["tipo"]=="Prazo")
        n_a = sum(1 for r in coord_rows if r["tipo"]=="Audiência")
        n_i = sum(1 for r in coord_rows if r["incons"])
        with cols_grid[i%3]:
            with st.container(border=True):
                st.markdown(f"**{coord_display}**")
                st.caption(f"{len(coord_rows)} ativ. · {n_p} prazos · {n_a} audiências · {n_i} inconsistências")
                buf = gerar_excel_coord(coord_key, active, coord_display)
                fname = coord_display[:30].replace(" ","_")
                st.download_button("📥 Excel",buf,f"IGSA_{fname}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_coord_{i}", use_container_width=True)

    st.markdown('<div class="section-title">Exportar Geral</div>', unsafe_allow_html=True)
    c1,c2 = st.columns(2)
    with c1:
        buf_all = exportar_xlsx_filtrado(active)
        st.download_button("📥 Exportar Tudo (Excel)", buf_all, "IGSA_Geral.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with c2:
        csv = exportar_csv(active)
        st.download_button("📄 Exportar Tudo (CSV)", csv, "IGSA_Geral.csv","text/csv", use_container_width=True)

def page_admin():
    render_header("Área Administrativa")
    st.markdown("""<div style="background:#FEF3C7;border:1px solid #F59E0B;border-radius:8px;
        padding:12px 16px;margin-bottom:16px;font-size:12px">
        ⚠️ <strong>Área restrita.</strong> Apenas a equipe da Controladoria Jurídica.
    </div>""", unsafe_allow_html=True)

    pub = load_published()
    if pub.get("publicado_em"):
        st.markdown(f'''<div class="published-banner">
            ✓ Dados publicados em {pub["publicado_em"]} · {fmt_num(pub["total"])} registros ·
            Referência: {pub.get("referencia","—")}
        </div>''', unsafe_allow_html=True)

    st.markdown('<div class="section-title">1. Carregar Planilha</div>', unsafe_allow_html=True)
    today = st.date_input("Data de referência", value=date.today())
    uploaded = st.file_uploader("Selecionar arquivo Excel (LegalOne)", type=["xlsx","xls"])
    if not uploaded: return

    st.markdown('<div class="section-title">2. Validação Automática</div>', unsafe_allow_html=True)
    with st.spinner("Processando e validando..."):
        try:
            df, headers = processar_planilha(uploaded, today)
            faltando = validar_estrutura(headers)
            if faltando: st.error(f"❌ Colunas ausentes: {faltando}"); return
            if len(df) < 5: st.error("❌ Arquivo com menos de 5 linhas."); return
            registros, alertas = construir_registros(df, today)
            st.success(f"✅ {fmt_num(len(df))} linhas · {fmt_num(len(registros))} até D-1 · {len(alertas)} inconsistências")
        except Exception as e:
            st.error(f"❌ Erro: {e}"); return

    cols = st.columns(5)
    stats = [
        ("Total D-1",fmt_num(len(registros)),"Atividades",""),
        ("Prazos",fmt_num(sum(1 for r in registros if r["tipo"]=="Prazo")),"Tipo Prazo","rose"),
        ("Audiências",fmt_num(sum(1 for r in registros if r["tipo"]=="Audiência")),"Tipo Audiência",""),
        ("Sem Coord.",fmt_num(sum(1 for r in registros if r["coord"]=="SEM COORDENADOR")),"Verificar","rose" if any(r["coord"]=="SEM COORDENADOR" for r in registros) else ""),
        ("Inconsistências",fmt_num(len(alertas)),"Registros","rose" if alertas else "green"),
    ]
    for i,(k,v,l,c) in enumerate(stats):
        with cols[i]: st.markdown(card(k,v,l,c), unsafe_allow_html=True)

    sem_coord = [r for r in registros if r["coord"]=="SEM COORDENADOR"]
    if sem_coord:
        st.markdown('<div class="section-title">⚠️ Sem Coordenador Mapeado</div>', unsafe_allow_html=True)
        sc_count = {}
        for r in sem_coord: sc_count[r["resp"]] = sc_count.get(r["resp"],0)+1
        st.dataframe(pd.DataFrame([{"Responsável":k,"Qtd":v} for k,v in sorted(sc_count.items(),key=lambda x:-x[1])]),
                     hide_index=True, use_container_width=True)
        st.warning("Estes responsáveis não aparecerão nos painéis de coordenador.")

    if alertas:
        st.markdown('<div class="section-title">⚠️ Inconsistências Detectadas</div>', unsafe_allow_html=True)
        st.dataframe(pd.DataFrame(alertas), hide_index=True, use_container_width=True, height=300)
        tipos_inc = {}
        for a in alertas:
            for t in a["Inconsistência"].split(";"):
                t=t.strip()
                if t: tipos_inc[t]=tipos_inc.get(t,0)+1
        for t,v in sorted(tipos_inc.items(),key=lambda x:-x[1]):
            st.markdown(f"- `{t}`: **{v}** ocorrências")
    else:
        st.success("✅ Nenhuma inconsistência detectada.")

    st.markdown('<div class="section-title">3. Publicar Dados</div>', unsafe_allow_html=True)
    if alertas: st.warning(f"⚠️ {len(alertas)} inconsistências. Revise ou prossiga assim mesmo.")

    c1,c2 = st.columns([2,1])
    with c1:
        versao = st.text_input("Identificação da carga",
            value=f"{today.strftime('%d/%m/%Y')} - Carga diária")
    with c2:
        github_token = st.text_input("GitHub Token (opcional)", type="password",
            help="Garante persistência dos dados no repositório")

    if st.button("🚀 PUBLICAR DADOS NO PAINEL", type="primary", use_container_width=True):
        with st.spinner("Publicando..."):
            pub_data = save_published(registros, versao, today.strftime("%d/%m/%Y"))
            msg = f"✅ {fmt_num(len(registros))} registros publicados em {pub_data['publicado_em']}."
            if github_token:
                with open(DATA_FILE, encoding="utf-8") as f: content = f.read()
                ok, err = push_to_github(github_token, "tatiikogan-beep/Gestao-Prazos-IGSA",
                    DATA_FILE, content, f"chore: publicar dados {versao}")
                msg += " GitHub ✓" if ok else f" ⚠️ GitHub falhou: {err}"
            st.success(msg)
            st.balloons()
            st.rerun()

# ── MAIN ───────────────────────────────────────────────────────────────────
def main():
    inject_css()
    page = render_sidebar()
    pub = load_published()
    registros = pub.get("registros", [])

    if page == "📊 Visão Geral": page_geral(registros)
    elif page == "👥 Por Coordenador": page_coordenador(registros)
    elif page == "👤 Por Responsável": page_responsavel(registros)
    elif page == "🔍 Auditoria": page_auditoria(registros)
    elif page == "📥 Exportação": page_exportacao(registros)
    elif page == "⚙️ Área Administrativa": page_admin()

if __name__ == "__main__":
    main()
